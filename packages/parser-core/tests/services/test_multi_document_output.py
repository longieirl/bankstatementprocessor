"""Tests for multi-document type output in various formats.

This module tests that document_type field is correctly included in
CSV, JSON, and Excel output files.
"""

from __future__ import annotations

import csv
import json
from pathlib import Path

import pytest

from bankstatements_core.patterns.strategies import (
    CSVOutputStrategy,
    JSONOutputStrategy,
)


@pytest.fixture
def sample_transactions():
    """Sample transactions with mixed document types."""
    return [
        {
            "Date": "01/12/2023",
            "Details": "Bank Transfer",
            "Debit_EUR": "500.00",
            "Credit_EUR": None,
            "Balance_EUR": "1500.00",
            "Filename": "bank_statement.pdf",
            "document_type": "bank_statement",
        },
        {
            "Date": "02/12/2023",
            "Details": "Card Purchase",
            "Debit_EUR": "100.00",
            "Credit_EUR": None,
            "Balance_EUR": "2000.00",
            "Filename": "credit_card.pdf",
            "document_type": "credit_card_statement",
        },
        {
            "Date": "03/12/2023",
            "Details": "Loan Payment",
            "Debit_EUR": "200.00",
            "Credit_EUR": None,
            "Balance_EUR": "1000.00",
            "Filename": "loan_statement.pdf",
            "document_type": "loan_statement",
        },
    ]


class TestCSVOutputWithDocumentType:
    """Test CSV output includes document_type column."""

    def test_csv_includes_document_type_column(self, sample_transactions, tmp_path):
        """Test that CSV output includes document_type as a column."""
        output_file = tmp_path / "output.csv"
        strategy = CSVOutputStrategy()
        column_names = list(sample_transactions[0].keys())

        strategy.write(sample_transactions, output_file, column_names)

        # Read CSV and verify structure
        with open(output_file, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            headers = reader.fieldnames

            # Verify document_type is in headers
            assert "document_type" in headers

            # Verify all rows have document_type
            rows = list(reader)
            assert len(rows) == 3
            for row in rows:
                assert "document_type" in row
                assert row["document_type"] in [
                    "bank_statement",
                    "credit_card_statement",
                    "loan_statement",
                ]

    def test_csv_preserves_document_type_values(self, sample_transactions, tmp_path):
        """Test that CSV output preserves correct document_type values."""
        output_file = tmp_path / "output.csv"
        strategy = CSVOutputStrategy()
        column_names = list(sample_transactions[0].keys())

        strategy.write(sample_transactions, output_file, column_names)

        # Read CSV and verify values
        with open(output_file, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            rows = list(reader)

            # Verify each row has correct document_type
            assert rows[0]["document_type"] == "bank_statement"
            assert rows[1]["document_type"] == "credit_card_statement"
            assert rows[2]["document_type"] == "loan_statement"


class TestJSONOutputWithDocumentType:
    """Test JSON output includes document_type field."""

    def test_json_includes_document_type_field(self, sample_transactions, tmp_path):
        """Test that JSON output includes document_type field."""
        output_file = tmp_path / "output.json"
        strategy = JSONOutputStrategy()
        column_names = list(sample_transactions[0].keys())

        strategy.write(sample_transactions, output_file, column_names)

        # Read JSON and verify structure
        with open(output_file, "r", encoding="utf-8") as f:
            data = json.load(f)

            assert len(data) == 3
            for item in data:
                assert "document_type" in item
                assert item["document_type"] in [
                    "bank_statement",
                    "credit_card_statement",
                    "loan_statement",
                ]

    def test_json_preserves_document_type_values(self, sample_transactions, tmp_path):
        """Test that JSON output preserves correct document_type values."""
        output_file = tmp_path / "output.json"
        strategy = JSONOutputStrategy()
        column_names = list(sample_transactions[0].keys())

        strategy.write(sample_transactions, output_file, column_names)

        # Read JSON and verify values
        with open(output_file, "r", encoding="utf-8") as f:
            data = json.load(f)

            assert data[0]["document_type"] == "bank_statement"
            assert data[1]["document_type"] == "credit_card_statement"
            assert data[2]["document_type"] == "loan_statement"


class TestExcelOutputWithDocumentType:
    """Test Excel output includes document_type column."""

    def test_excel_includes_document_type_column(self, sample_transactions, tmp_path):
        """Test that Excel output includes document_type column."""
        pytest.importorskip("openpyxl", reason="openpyxl not installed (PAID tier)")

        from bankstatements_core.patterns.strategies import ExcelOutputStrategy

        output_file = tmp_path / "output.xlsx"
        strategy = ExcelOutputStrategy()
        column_names = list(sample_transactions[0].keys())

        strategy.write(sample_transactions, output_file, column_names)

        # Read Excel and verify structure
        import openpyxl

        wb = openpyxl.load_workbook(output_file)
        ws = wb.active

        # Get headers (first row)
        headers = [cell.value for cell in ws[1]]
        assert "document_type" in headers

        # Verify all data rows have document_type
        doc_type_col_idx = headers.index("document_type") + 1
        for row_idx in range(2, ws.max_row + 1):  # Skip header row
            doc_type_value = ws.cell(row_idx, doc_type_col_idx).value
            assert doc_type_value in [
                "bank_statement",
                "credit_card_statement",
                "loan_statement",
            ]

    def test_excel_preserves_document_type_values(self, sample_transactions, tmp_path):
        """Test that Excel output preserves correct document_type values."""
        pytest.importorskip("openpyxl", reason="openpyxl not installed (PAID tier)")

        from bankstatements_core.patterns.strategies import ExcelOutputStrategy

        output_file = tmp_path / "output.xlsx"
        strategy = ExcelOutputStrategy()
        column_names = list(sample_transactions[0].keys())

        strategy.write(sample_transactions, output_file, column_names)

        # Read Excel and verify values
        import openpyxl

        wb = openpyxl.load_workbook(output_file)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        doc_type_col_idx = headers.index("document_type") + 1

        # Verify each row has correct document_type
        assert ws.cell(2, doc_type_col_idx).value == "bank_statement"
        assert ws.cell(3, doc_type_col_idx).value == "credit_card_statement"
        assert ws.cell(4, doc_type_col_idx).value == "loan_statement"
