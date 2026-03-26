"""
Strategy Pattern implementations for interchangeable algorithms.

This module provides abstract interfaces and concrete implementations for
algorithms that can be swapped at runtime (duplicate detection, output
formatting, etc.).
"""

from __future__ import annotations

import json
import logging
from abc import ABC, abstractmethod
from pathlib import Path
from typing import TYPE_CHECKING, Any

import pandas as pd

from bankstatements_core.utils import to_float

if TYPE_CHECKING:
    from bankstatements_core.entitlements import Entitlements  # noqa: F401

logger = logging.getLogger(__name__)


class DuplicateDetectionStrategy(ABC):
    """Abstract strategy for detecting duplicate transactions."""

    @abstractmethod
    def create_key(self, transaction: dict) -> str:
        """
        Create a unique key for a transaction.

        Duplicate transactions should produce the same key.

        Args:
            transaction: Transaction dictionary

        Returns:
            String key that uniquely identifies the transaction
        """
        pass

    def detect_duplicates(
        self, transactions: list[dict]
    ) -> tuple[list[dict], list[dict]]:
        """
        Detect duplicates in a list of transactions.

        Args:
            transactions: List of transaction dictionaries

        Returns:
            Tuple of (unique_transactions, duplicate_transactions)
        """
        unique_rows = []
        duplicate_rows = []
        transaction_files: dict[str, str] = {}

        for row in transactions:
            transaction_key = self.create_key(row)
            current_filename = row.get("Filename", "")

            if transaction_key in transaction_files:
                # Same transaction details but different file = duplicate
                if transaction_files[transaction_key] != current_filename:
                    duplicate_rows.append(row)
                else:
                    # Same file, same transaction - keep it
                    unique_rows.append(row)
            else:
                # First time seeing this transaction
                transaction_files[transaction_key] = current_filename
                unique_rows.append(row)

        return unique_rows, duplicate_rows


class AllFieldsDuplicateStrategy(DuplicateDetectionStrategy):
    """
    Duplicate detection strategy that matches on all relevant fields.

    This is the strictest strategy - transactions must match on date,
    details, and all monetary fields to be considered duplicates.
    """

    def create_key(self, transaction: dict) -> str:
        """Create key from date, details, and all monetary columns."""
        # Extract date and details
        date = transaction.get("Date", "")
        details = transaction.get("Details", "")

        # Extract monetary values (any column with €, £, $, etc.)
        amounts = []
        for key, value in transaction.items():
            if key not in ["Date", "Details", "Filename"] and value:
                # Try to parse as currency
                amount = to_float(str(value))
                if amount is not None:
                    amounts.append(f"{key}:{amount}")

        # Combine into key
        amounts_str = "|".join(sorted(amounts))
        return f"{date}:{details}:{amounts_str}"


class DateAmountDuplicateStrategy(DuplicateDetectionStrategy):
    """
    Lenient duplicate detection strategy that only matches on date and total amount.

    This strategy is useful when transaction descriptions may vary slightly
    between statements, but the date and amount should match.
    """

    def create_key(self, transaction: dict) -> str:
        """Create key from date and sum of all amounts."""
        date = transaction.get("Date", "")

        # Sum all monetary values
        total = 0.0
        for key, value in transaction.items():
            if key not in ["Date", "Details", "Filename"] and value:
                amount = to_float(str(value))
                if amount is not None:
                    total += abs(amount)  # Use absolute value

        return f"{date}:{total:.2f}"


class CreditCardDuplicateStrategy(DuplicateDetectionStrategy):
    """Credit card aware duplicate detection strategy.

    Improvements over basic strategies:
    - Considers transaction_type (don't match purchase with payment)
    - Uses date + amount + transaction_type for matching
    - More tolerant of description variations (merchants often vary)

    This strategy is specifically designed for credit card statements where:
    - Multiple purchases of the same amount on the same day are common
    - Payments and purchases with the same amount should not be duplicates
    - Transaction type provides critical disambiguation
    """

    def create_key(self, transaction: dict) -> str:
        """Create composite key: date:transaction_type:amount.

        Args:
            transaction: Transaction dictionary

        Returns:
            Unique key combining date, transaction type, and amount
        """
        date = transaction.get("Date", "")
        transaction_type = transaction.get("transaction_type", "other")

        # Get primary amount (prefer Debit, then Credit)
        amount = None
        if transaction.get("Debit_AMT"):
            amount = to_float(str(transaction.get("Debit_AMT")))
        elif transaction.get("Credit_AMT"):
            amount = to_float(str(transaction.get("Credit_AMT")))

        amount_str = f"{amount:.2f}" if amount is not None else "0.00"

        # Key: date + type + amount (ignoring description variations)
        # This allows a $50 purchase and $50 payment on the same day
        # to be treated as different transactions
        return f"{date}:{transaction_type}:{amount_str}"


class OutputFormatStrategy(ABC):
    """
    Abstract strategy for writing transactions in different formats.

    Uses Template Method pattern to define the overall write algorithm
    while allowing subclasses to customize specific steps.
    """

    def write(
        self,
        transactions: list[dict],
        file_path: Path,
        column_names: list[str],
        **kwargs: Any,
    ) -> None:
        """
        Template method that defines the algorithm for writing transactions.

        This method orchestrates the write process:
        1. Extract common parameters (include_totals, totals_columns)
        2. Convert data to appropriate format (DataFrame, dict, etc.)
        3. Write the main data
        4. Optionally write totals (if supported)

        Args:
            transactions: List of transaction dictionaries
            file_path: Path to write the output file
            column_names: Ordered list of column names
            **kwargs: Additional format-specific options
        """
        # Step 1: Extract common parameters
        include_totals = kwargs.get("include_totals", False)
        totals_columns = kwargs.get("totals_columns", [])

        # Step 2: Convert data to appropriate format
        data = self._prepare_data(transactions, column_names)

        # Step 3: Write main data
        self._write_data(data, file_path, column_names, **kwargs)

        # Step 4: Optionally write totals (hook method - can be overridden)
        if include_totals and totals_columns and self._supports_totals():
            totals_row = kwargs.get("totals_row", [])
            if totals_row:
                self._write_totals(file_path, totals_row)

    @abstractmethod
    def _prepare_data(self, transactions: list[dict], column_names: list[str]) -> Any:
        """
        Prepare transactions data in format-specific structure.

        Args:
            transactions: List of transaction dictionaries
            column_names: Ordered list of column names

        Returns:
            Format-specific data structure (DataFrame, list, etc.)
        """
        pass

    @abstractmethod
    def _write_data(
        self, data: Any, file_path: Path, column_names: list[str], **kwargs: Any
    ) -> None:
        """
        Write the main transaction data to file.

        Args:
            data: Prepared data from _prepare_data()
            file_path: Path to write the output file
            column_names: Ordered list of column names
            **kwargs: Additional format-specific options
        """
        pass

    def _supports_totals(self) -> bool:
        """
        Hook method: Does this format support totals?

        Returns:
            True if format supports totals, False otherwise
        """
        return False

    def _write_totals(  # noqa: B027
        self,
        file_path: Path,
        totals_row: list[str],
    ) -> None:
        """
        Hook method: Write pre-calculated totals row (optional, override if supported).

        This is intentionally not abstract - it's a hook that subclasses
        can override if they support totals. Not abstractmethod because
        not all output strategies need to implement this.

        Calculation is now done in the service layer before calling this method,
        so strategies only handle formatting/writing, not calculation.

        Args:
            file_path: Path to output file
            totals_row: Pre-calculated totals row values
        """
        pass


class CSVOutputStrategy(OutputFormatStrategy):
    """Strategy for writing transactions as CSV files."""

    def _prepare_data(
        self, transactions: list[dict], column_names: list[str]
    ) -> pd.DataFrame:
        """Prepare data as DataFrame for CSV output."""
        return pd.DataFrame(transactions, columns=column_names)

    def _write_data(
        self,
        data: pd.DataFrame,
        file_path: Path,
        column_names: list[str],
        **kwargs: Any,
    ) -> None:
        """Write DataFrame to CSV file."""
        data.to_csv(file_path, index=False, encoding="utf-8")

    def _supports_totals(self) -> bool:
        """CSV format supports totals."""
        return True

    def _write_totals(
        self,
        file_path: Path,
        totals_row: list[str],
    ) -> None:
        """Append pre-calculated totals row to CSV file."""
        # Append totals row to file
        with open(file_path, "a", newline="", encoding="utf-8") as csvfile:
            csvfile.write("\n")
            csvfile.write(",".join(f'"{value}"' for value in totals_row))
            csvfile.write("\n")


class JSONOutputStrategy(OutputFormatStrategy):
    """Strategy for writing transactions as JSON files."""

    def _prepare_data(
        self, transactions: list[dict], column_names: list[str]
    ) -> list[dict]:
        """Prepare data as list for JSON output."""
        return transactions

    def _write_data(
        self, data: list[dict], file_path: Path, column_names: list[str], **kwargs: Any
    ) -> None:
        """Write transactions list to JSON file."""
        file_path.write_text(json.dumps(data, indent=2), encoding="utf-8")

    # JSON doesn't support totals, so _supports_totals() returns False by default


class ExcelOutputStrategy(OutputFormatStrategy):
    """Strategy for writing transactions as Excel files (.xlsx)."""

    def _prepare_data(
        self, transactions: list[dict], column_names: list[str]
    ) -> pd.DataFrame:
        """Prepare data as DataFrame with numeric conversion for Excel."""
        df = pd.DataFrame(transactions, columns=column_names)
        return self._convert_numeric_columns(df)

    def _write_data(
        self,
        data: pd.DataFrame,
        file_path: Path,
        column_names: list[str],
        **kwargs: Any,
    ) -> None:
        """Write DataFrame to Excel file with formatting."""
        # Store for totals writing (needs writer context)
        self._current_writer = None
        self._current_file_path = file_path

        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            self._current_writer = writer
            data.to_excel(writer, sheet_name="Transactions", index=False)
            self._apply_number_formatting(writer, data, column_names)
            # Keep writer open for totals if needed

    def _supports_totals(self) -> bool:
        """Excel format supports totals."""
        return True

    def _write_totals(
        self,
        file_path: Path,
        totals_row: list[str],
    ) -> None:
        """Append pre-calculated totals row to Excel file."""
        from openpyxl import load_workbook

        # Re-open workbook to append totals
        workbook = load_workbook(file_path)
        worksheet = workbook["Transactions"]

        # Determine row number (+1 for header, +1 for empty row, +1 for totals)
        row_num = worksheet.max_row + 2

        # Write totals row
        for col_idx, value in enumerate(totals_row, start=1):
            if value:  # Only write non-empty values
                # Convert string to numeric if it's not "TOTAL"
                try:
                    cell_value = float(value) if value != "TOTAL" else value
                except ValueError:
                    cell_value = value
                worksheet.cell(row=row_num, column=col_idx, value=cell_value)

        # Save workbook
        workbook.save(file_path)

    def _is_numeric_column(self, column_name: str) -> bool:
        """
        Check if a column should be treated as numeric based on its name.

        Args:
            column_name: The name of the column

        Returns:
            True if the column is numeric, False otherwise
        """
        # Patterns that indicate numeric columns
        numeric_patterns = [
            "debit",
            "credit",
            "balance",
            "amount",
            "total",
            "price",
            "cost",
            "fee",
            "charge",
            "€",
            "$",
            "£",
            "¥",
        ]

        column_lower = column_name.lower()
        return any(pattern in column_lower for pattern in numeric_patterns)

    def _convert_numeric_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Convert numeric columns from strings to actual numbers.

        Args:
            df: DataFrame with potential string numeric values

        Returns:
            DataFrame with numeric columns converted to float type
        """
        df_copy = df.copy()

        for col in df_copy.columns:
            if self._is_numeric_column(col):
                # Convert using the to_float utility
                def convert_to_numeric(x: Any) -> float | None:
                    if pd.notna(x):
                        x_str = str(x).strip()
                        if x_str:
                            return to_float(x_str)
                    return None

                df_copy[col] = df_copy[col].apply(convert_to_numeric)

        return df_copy

    def _apply_number_formatting(
        self, writer: pd.ExcelWriter, df: pd.DataFrame, column_names: list[str]
    ) -> None:
        """
        Apply number formatting to numeric columns in Excel.

        Args:
            writer: ExcelWriter instance
            df: DataFrame containing the data
            column_names: Ordered list of column names
        """
        from openpyxl.styles import numbers

        worksheet = writer.sheets["Transactions"]

        # Apply number format to numeric columns
        for col_idx, col_name in enumerate(column_names, start=1):
            if self._is_numeric_column(col_name):
                # Apply currency format with 2 decimal places: #,##0.00
                for row_idx in range(2, len(df) + 2):  # Start from row 2 (after header)
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1


def create_output_strategy(
    format_name: str, entitlements: "Entitlements"
) -> OutputFormatStrategy:
    """
    Create output strategy with entitlement enforcement.

    This factory function enforces tier-based access control for output formats.
    FREE tier users can only output CSV, while PAID tier users can output
    CSV, JSON, and Excel formats.

    Args:
        format_name: Output format ("csv", "json", or "excel"/"xlsx")
        entitlements: Entitlements to enforce

    Returns:
        Appropriate OutputFormatStrategy instance

    Raises:
        EntitlementError: If format is not allowed for the tier
        ValueError: If format_name is not recognized

    Examples:
        >>> from bankstatements_core.entitlements import Entitlements
        >>> # FREE tier - only CSV allowed
        >>> ent = Entitlements.free_tier()
        >>> strategy = create_output_strategy("csv", ent)  # OK
        >>> strategy = create_output_strategy("json", ent)  # Raises EntitlementError
        >>> # PAID tier - all formats allowed
        >>> ent = Entitlements.paid_tier()
        >>> strategy = create_output_strategy("json", ent)  # OK
    """
    from bankstatements_core.entitlements import Entitlements  # noqa: F401

    # Normalize format name
    format_lower = format_name.lower()

    # Normalize "excel" to "xlsx" for entitlement checking
    # (the strategy factory accepts both, but entitlements use "xlsx")
    entitlement_format = "xlsx" if format_lower == "excel" else format_lower

    # Enforce entitlements BEFORE creating strategy
    entitlements.check_output_format(entitlement_format)

    logger.info(
        f"Creating {format_lower.upper()} output strategy ({entitlements.tier} tier)"
    )

    # Map format names to strategy classes
    # Note: both "excel" and "xlsx" map to ExcelOutputStrategy
    if format_lower == "csv":
        return CSVOutputStrategy()
    elif format_lower == "json":
        return JSONOutputStrategy()
    elif format_lower in ("excel", "xlsx"):
        return ExcelOutputStrategy()
    else:
        raise ValueError(
            f"Unknown output format: {format_name}. "
            f"Valid formats: csv, json, excel, xlsx"
        )
